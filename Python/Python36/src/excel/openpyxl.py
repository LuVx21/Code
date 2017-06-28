# coding=utf-8

# openpyxl可以对excel文件进行读写操作


# from openpyxl import load_workbook
# from openpyxl import workbook
# from openpyxl.writer.excel import ExcelWriter
from openpyxl.reader import excel

workbook_ = load_workbook(u"1.xlsx")
sheetnames =workbook_.get_sheet_names() #获得表单名字
print(sheetnames)
'''
sheet = workbook_.get_sheet_by_name(sheetnames[0])
print(sheet.cell(row=3,column=3).value)
sheet['A1'] = '47' 
workbook_.save(u"新歌检索失败1477881109469_new.xlsx")  
wb = Workbook()
ws = wb.active
ws['A1'] = 4
wb.save("新歌检索失败.xlsx")
'''