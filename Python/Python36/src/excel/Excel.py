# coding=utf-8

# xlrd xlwt xlutils openpyxl xlsxwriter
import xlrd
import xlwt
import xlutils
import openpyxl
import xlsxwriter




 

# 4.openpyxl可以对excel文件进行读写操作

from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter 

 

workbook_ = load_workbook(u"新歌检索失败1477881109469.xlsx")

sheetnames =workbook_.get_sheet_names() #获得表单名字

print sheetnames

sheet = workbook_.get_sheet_by_name(sheetnames[0])

print sheet.cell(row=3,column=3).value

sheet['A1'] = '47' 

workbook_.save(u"新歌检索失败1477881109469_new.xlsx")  

wb = Workbook()

ws = wb.active

ws['A1'] = 4

wb.save("新歌检索失败.xlsx") 

'''  

# 5.xlsxwriter可以写excel文件并加上图表

import xlsxwriter

def get_chart(series):

    chart = workbook.add_chart({'type': 'line'})

    for ses in series:

        name = ses["name"]

        values = ses["values"]

        chart.add_series({ 

            'name': name,

            'categories': 'A2:A10',

            'values':values

        })  

    chart.set_size({'width': 700, 'height': 350}) 

    return chart

 

if __name__ == '__main__':

    workbook = xlsxwriter.Workbook(u'H5应用中心关键数据及趋势.xlsx') 

    worksheet = workbook.add_worksheet(u"每日PV,UV")

    headings = ['日期', '平均值']

    worksheet.write_row('A1', headings)

    index=0

    for row in range(1,10):

        for com in [0,1]:

            worksheet.write(row,com,index)

            index+=1  

    series = [{"name":"平均值","values":"B2:B10"}]

    chart = get_chart(series)

    chart.set_title ({'name': '每日页面分享数据'})  

    worksheet.insert_chart('H7', chart)

    workbook.close()

 

 

openpyxl 
'''